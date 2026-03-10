# -*- coding: utf-8 -*-
"""
TÍNH CHỈ SỐ PHÁI SINH THỊ TRƯỜNG BĐS — PHÍA TÂY HÀ NỘI
===========================================================
Đọc dữ liệu từ crawl_chi_so_thi_truong.py → tính các chỉ số phái sinh:
  1. Giá Gap (so sánh giá giữa khu vực)
  2. Cung-Cầu Ratio (ước tính)
  3. MFV — Money Flow Velocity (tốc độ dòng tiền)
  4. Cycle Index (nhiệt kế thị trường 0-100)
  5. MFSI — Money Flow Shift Index (dòng tiền dịch chuyển)
  6. Heat Score (chỉ số nóng tổng hợp)

Cách chạy:
  python tinh_chi_so_phai_sinh.py
  python tinh_chi_so_phai_sinh.py --only hoai-duc,nam-tu-liem
"""

import sys
import io
import csv
import json
import argparse
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import os
OUTPUT_DIR = Path(os.environ["BDS_DATA_DIR"]) if "BDS_DATA_DIR" in os.environ else Path(__file__).parent

# Defensive utils (DC01-DC07)
try:
    from defensive_utils import (
        safe_append_csv, backup_csv, check_config_freshness,
        log_warning, log_error, log_success,
    )
    HAS_DEFENSIVE = True
except ImportError:
    HAS_DEFENSIVE = False


# ============================================================
# ĐỌC DỮ LIỆU ĐẦU VÀO
# ============================================================

def load_crawl_data(filepath: Path) -> dict:
    """Đọc dữ liệu mới nhất từ lich_su_chi_so.csv — lấy dòng cuối cùng mỗi khu vực."""
    data = {}
    if not filepath.exists():
        print(f"!! Không tìm thấy file: {filepath}")
        return data

    # DC04: Backup trước khi đọc
    if HAS_DEFENSIVE:
        backup_csv(filepath)

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Khu vực", "").strip()
            if not name:
                continue
            # Luôn ghi đè → cuối cùng sẽ giữ dòng mới nhất
            data[name] = {
                "ngay": row.get("Ngày", ""),
                "total_ban": _float(row.get("Tổng tin bán")),
                "views_ban": _float(row.get("Lượt xem khu vực")),
                "total_thue": _float(row.get("Tổng tin thuê")),
                "distress_count": _float(row.get("Tin cắt lỗ")),
                "distress_ratio": _float(row.get("% Cắt lỗ")),
                "today_count": _float(row.get("Đăng hôm nay")),
                "pct_today": _float(row.get("% Hôm nay")),
                "avg_price_per_m2": _float(row.get("Giá bán TB (tr/m²)")),
                "avg_rent_price": _float(row.get("Giá thuê TB (tr/th)")),
                "rental_yield": _float(row.get("Rental Yield (%)")),
            }
    return data


def load_config(filepath: Path) -> dict:
    """Đọc config thủ công (views/tin, giá YoY, v.v.)"""
    if not filepath.exists():
        print(f"!! Không tìm thấy config: {filepath}")
        return {}
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def _float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return None


def load_previous_totals(history_file: Path) -> dict:
    """Đọc ngày gần nhất trước hôm nay từ lịch sử → lấy Tổng tin cho mỗi KV."""
    if not history_file.exists():
        return {}

    # Đọc tất cả rows, nhóm theo ngày
    by_date = {}
    with open(history_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            date = row.get("Ngày", "").strip()
            name = row.get("Khu vực", "").strip()
            if date and name:
                if date not in by_date:
                    by_date[date] = {}
                by_date[date][name] = _float(row.get("Tổng tin"))

    # Sắp xếp ngày → lấy ngày gần nhất (trước ngày cuối cùng)
    sorted_dates = sorted(by_date.keys())
    if len(sorted_dates) < 2:
        return {}  # Chưa đủ 2 ngày data

    prev_date = sorted_dates[-2]  # Ngày trước hôm nay
    print(f"  → Dữ liệu ngày trước: {prev_date} ({len(by_date[prev_date])} KV)")
    return by_date[prev_date]


# ============================================================
# TÍNH CHỈ SỐ PHÁI SINH
# ============================================================

def calculate_derived(crawl_data: dict, config: dict) -> dict:
    """Tính tất cả chỉ số phái sinh cho từng khu vực."""
    views_config = config.get("views_per_tin", {})
    yoy_config = config.get("gia_yoy_pct", {})
    # Benchmark = giá trung bình TẤT CẢ khu vực
    all_prices = [d.get("avg_price_per_m2") for d in crawl_data.values()
                  if d.get("avg_price_per_m2") and d["avg_price_per_m2"] > 0]
    benchmark_price = sum(all_prices) / len(all_prices) if all_prices else None
    if benchmark_price:
        print(f"  → Giá benchmark (TB {len(all_prices)} quận): {benchmark_price:.1f} tr/m²")

    # Load dữ liệu ngày trước để tính Tỷ lệ thay đổi Cung
    history_file = OUTPUT_DIR / "lich_su_chi_so.csv"
    prev_totals = load_previous_totals(history_file)
    has_prev = len(prev_totals) > 0
    if not has_prev:
        print("  → Chưa có dữ liệu ngày trước → dùng fallback cho supply_change")

    results = {}
    for name, d in crawl_data.items():
        gia_yoy = yoy_config.get(name, 0.0)
        price = d.get("avg_price_per_m2") or 0
        total_ban = d.get("total_ban") or 1
        pct_today = d.get("pct_today") or 0
        distress_ratio = d.get("distress_ratio") or 0
        rental_yield = d.get("rental_yield") or 0

        # Views: ưu tiên dữ liệu crawl thật, fallback sang config
        real_views = d.get("views_ban")
        if real_views and total_ban > 0:
            total_views = int(real_views)
            views_tin = round(real_views / total_ban, 1)
        else:
            views_tin = views_config.get(name, 5.0)
            total_views = round(views_tin * total_ban)

        # 1. GIÁ GAP vs benchmark
        gia_gap_pct = None
        if benchmark_price and price and benchmark_price > 0:
            gia_gap_pct = (price - benchmark_price) / benchmark_price * 100

        # 2. CUNG-CẦU RATIO (ước tính)
        cung_cau = views_tin / 5.0  # normalize: 5 views = cân bằng (1.0)

        # 3. TỶ LỆ THAY ĐỔI CUNG (thay thế absorption)
        # supply_change_pct > 0 = cung tăng (xấu), < 0 = cung giảm (tốt)
        prev_total = prev_totals.get(name)
        if has_prev and prev_total and prev_total > 0:
            supply_change_pct = (total_ban - prev_total) / prev_total * 100
        else:
            supply_change_pct = 0  # Chưa có data → mặc định trung tính

        # 4. MFV — Money Flow Velocity
        # MFV = Tổng tin × Views/Tin × (1 + giá YoY/100)
        # Đo "tổng lượng quan tâm × động lực giá" — chỉ dùng data crawl thật
        mfv = total_ban * views_tin * (1 + gia_yoy / 100) / 100

        # 5. CYCLE INDEX (0-100)
        cycle = _calc_cycle(
            distress_ratio=distress_ratio,
            cung_cau=cung_cau,
            gia_yoy=gia_yoy,
            supply_change_pct=supply_change_pct,
            views_tin=views_tin,
        )

        # 6. MFSI — Money Flow Shift Index
        mfsi = _calc_mfsi(
            supply_change_pct=supply_change_pct,
            gia_gap_pct=gia_gap_pct,
            views_tin=views_tin,
            distress_ratio=distress_ratio,
        )

        # 7. HEAT SCORE
        heat = _calc_heat(
            mfv=mfv,
            views_tin=views_tin,
            cung_cau=cung_cau,
            gia_yoy=gia_yoy,
            distress_ratio=distress_ratio,
        )

        # Tính tổng views ước tính
        total_views = round(views_tin * total_ban)

        results[name] = {
            "views_tin": views_tin,
            "total_views": total_views,
            "gia_yoy": gia_yoy,
            "gia_gap_pct": gia_gap_pct,
            "cung_cau": cung_cau,
            "mfv": round(mfv, 1),
            "cycle": round(cycle, 0),
            "mfsi": round(mfsi, 1),
            "heat_score": round(heat, 0),
            # Dữ liệu gốc
            "price": price,
            "total_ban": total_ban,
            "distress_ratio": distress_ratio,
            "pct_today": pct_today,
            "rental_yield": rental_yield,
            "supply_change_pct": round(supply_change_pct, 1),
        }

    return results


def _calc_cycle(distress_ratio, cung_cau, gia_yoy, supply_change_pct, views_tin) -> float:
    """
    Cycle Index (0-100): Đo vị trí trong chu kỳ BĐS.
    Công thức tổng hợp 5 yếu tố:
    - Cắt lỗ thấp → điểm cao (thị trường khỏe)
    - Cung-cầu cao → điểm cao (cầu > cung)
    - Giá tăng YoY → điểm cao
    - Cung giảm → điểm cao (tin bị hấp thụ)
    - Views/Tin cao → điểm cao (quan tâm nhiều)
    """
    # Điểm cắt lỗ (ngược: thấp = tốt): 0-20 điểm
    if distress_ratio <= 1:
        d_score = 20
    elif distress_ratio <= 3:
        d_score = 16
    elif distress_ratio <= 5:
        d_score = 12
    elif distress_ratio <= 10:
        d_score = 6
    else:
        d_score = 2

    # Điểm cung-cầu: 0-20 điểm
    cc_score = min(20, max(0, cung_cau * 10))

    # Điểm giá tăng YoY: 0-25 điểm
    if gia_yoy >= 30:
        y_score = 25
    elif gia_yoy >= 20:
        y_score = 20
    elif gia_yoy >= 10:
        y_score = 15
    elif gia_yoy >= 0:
        y_score = 8
    else:
        y_score = 2

    # Điểm thay đổi cung (cung GIẢM = tốt): 0-15 điểm
    # supply_change_pct < 0 = cung giảm = thị trường hấp thụ tốt
    if supply_change_pct <= -10:
        s_score = 15  # Cung giảm mạnh
    elif supply_change_pct <= -5:
        s_score = 12
    elif supply_change_pct <= -1:
        s_score = 9
    elif supply_change_pct <= 1:
        s_score = 6   # Cung ổn định
    elif supply_change_pct <= 5:
        s_score = 3
    else:
        s_score = 1   # Cung tăng mạnh = xấu

    # Điểm views/tin: 0-20 điểm
    v_score = min(20, max(0, views_tin * 2))

    total = d_score + cc_score + y_score + s_score + v_score
    return min(100, max(0, total))


def _calc_mfsi(supply_change_pct, gia_gap_pct, views_tin, distress_ratio) -> float:
    """
    MFSI: Money Flow Shift Index.
    Đo mức hấp dẫn của dòng tiền vào khu vực.
    Công thức: Biến động cung (25) + Giá gap (50) + Tâm lý (25)
    """
    # Biến động cung (25 điểm max)
    if supply_change_pct <= -10:
        liq = 18
    elif supply_change_pct <= -5:
        liq = 16
    elif supply_change_pct <= -1:
        liq = 14
    elif supply_change_pct <= 1:
        liq = 12  # Ổn định
    elif supply_change_pct <= 5:
        liq = 6
    else:
        liq = 2   # Cung tăng mạnh

    # Giá gap vs benchmark TB 16 quận (50 điểm max)
    if gia_gap_pct is None:
        gap = 25
    elif gia_gap_pct <= -30:
        gap = 50
    elif gia_gap_pct <= -15:
        gap = 40
    elif gia_gap_pct <= 0:
        gap = 30
    elif gia_gap_pct <= 15:
        gap = 15
    else:
        gap = 5

    # Tâm lý thị trường (25 điểm max)
    views_score = min(15, views_tin * 1.5)                # max 15đ
    distress_score = min(10, max(0, 10 - distress_ratio)) # max 10đ
    sent = views_score + distress_score

    return liq + gap + sent


def _calc_heat(mfv, views_tin, cung_cau, gia_yoy, distress_ratio) -> float:
    """
    Heat Score: Chỉ số nóng tổng hợp.
    Công thức: MFV (40%) + Views (25%) + Cung-cầu (15%) + YoY (15%) + Sức khỏe (5%)
    """
    # Normalize MFV (0-400 range → 0-100)
    mfv_score = min(100, mfv / 4) * 0.4

    # Views/Tin (0-15 range → 0-100)
    views_score = min(100, views_tin / 15 * 100) * 0.25

    # Cung-cầu (0-2 range → 0-100)
    cc_score = min(100, cung_cau / 2 * 100) * 0.15

    # Giá tăng YoY (0-50% range → 0-100)
    yoy_score = min(100, gia_yoy / 50 * 100) * 0.15

    # Sức khỏe (cắt lỗ thấp = khỏe): 0-100
    health_score = max(0, 100 - distress_ratio * 5) * 0.05

    total = mfv_score + views_score + cc_score + yoy_score + health_score
    return total * 10  # Scale lên 0-1000 để dễ so sánh


# ============================================================
# OUTPUT
# ============================================================

def fmt(val, decimals=1, suffix=""):
    if val is None:
        return "N/A"
    return f"{val:,.{decimals}f}{suffix}"


def cycle_label(c):
    if c is None: return "N/A"
    if c >= 85: return "🍂 MÙA THU (quá nóng)"
    if c >= 60: return "☀️ MÙA HÈ (đang tăng)"
    if c >= 30: return "🌱 MÙA XUÂN (bắt đầu)"
    return "🥶 MÙA ĐÔNG (đáy)"


def cycle_emoji(c):
    if c is None: return ""
    if c >= 85: return "🔴"
    if c >= 60: return "🟢"
    if c >= 30: return "🟡"
    return "⚫"


def cc_label(cc):
    if cc is None: return "N/A"
    if cc >= 1.5: return "🔥 Cầu >> Cung"
    if cc >= 1.0: return "✅ Cầu > Cung"
    if cc >= 0.7: return "🟡 Cân bằng"
    return "❌ Cung > Cầu"


def generate_report(results: dict, crawl_data: dict, config: dict) -> str:
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
    config_date = config.get("_ngay_cap_nhat", "N/A")
    benchmark = config.get("benchmark_khu_vuc", "Nam Từ Liêm")

    # Sắp xếp theo Heat Score giảm dần
    sorted_results = sorted(results.items(), key=lambda x: x[1].get("heat_score", 0), reverse=True)

    lines = [
        "# 🔥 CHỈ SỐ PHÁI SINH THỊ TRƯỜNG — BĐS HÀ NỘI",
        "",
        f"> **Ngày tính:** {timestamp}",
        f"> **Dữ liệu crawl:** batdongsan.com.vn | **Config (Views/YoY):** {config_date}",
        f"> **Benchmark:** {benchmark}",
        "",
        "---",
        "",
        "## I. BẢNG TỔNG HỢP (sắp xếp theo Heat Score)",
        "",
        "| # | Khu vực | Tổng tin | Tổng views | Views/Tin | Cycle | Giai đoạn | Heat | MFV | Cung-Cầu | Giá gap | MFSI | Giá (tr/m²) |",
        "|:-:|---------|:-------:|:----------:|:---------:|:-----:|-----------|:----:|:---:|:--------:|:-------:|:----:|:-----------:|",
    ]

    for i, (name, r) in enumerate(sorted_results, 1):
        gap_str = f"{r['gia_gap_pct']:+.0f}%" if r['gia_gap_pct'] is not None else "—"
        lines.append(
            f"| {i} | **{name}** "
            f"| {fmt(r['total_ban'], 0)} "
            f"| {fmt(r['total_views'], 0)} "
            f"| {fmt(r['views_tin'], 1)} "
            f"| {cycle_emoji(r['cycle'])} {fmt(r['cycle'], 0)} "
            f"| {cycle_label(r['cycle'])} "
            f"| **{fmt(r['heat_score'], 0)}** "
            f"| {fmt(r['mfv'], 0)} "
            f"| {fmt(r['cung_cau'], 2)} {cc_label(r['cung_cau'])} "
            f"| {gap_str} "
            f"| {fmt(r['mfsi'], 1)} "
            f"| {fmt(r['price'], 1)} |"
        )

    # Đèn giao thông
    lines.extend([
        "", "---", "",
        "## II. ĐÈN GIAO THÔNG ĐẦU TƯ",
        "",
        "| Khu vực | Đèn | Cycle | MFSI | Khuyến nghị |",
        "|---------|:----:|:-----:|:----:|-------------|",
    ])

    for name, r in sorted_results:
        cycle = r.get("cycle", 0)
        mfsi = r.get("mfsi", 0)
        if cycle >= 85:
            den = "🔴"
            kn = "THẬN TRỌNG — gần đỉnh, rủi ro cao"
        elif cycle >= 60 and mfsi >= 50:
            den = "🟢"
            kn = "CƠ HỘI — đang tăng, dòng tiền vào"
        elif cycle >= 60 and mfsi < 50:
            den = "🟡"
            kn = "CÂN NHẮC — tăng nhưng dòng tiền yếu"
        elif cycle >= 30:
            den = "⚫→🟢"
            kn = "SỚM — cần kiên nhẫn 3-5 năm"
        else:
            den = "⚫"
            kn = "CHỜ — chưa có tín hiệu"
        lines.append(f"| **{name}** | {den} | {fmt(cycle, 0)} | {fmt(mfsi, 1)} | {kn} |")

    # Top picks
    lines.extend([
        "", "---", "",
        "## III. TOP 3 KHU VỰC HẤP DẪN NHẤT",
        "",
    ])

    for i, (name, r) in enumerate(sorted_results[:3], 1):
        lines.extend([
            f"### {i}. {name} — Heat Score {fmt(r['heat_score'], 0)}",
            "",
            f"- **Cycle:** {fmt(r['cycle'], 0)} ({cycle_label(r['cycle'])})",
            f"- **Tổng tin:** {fmt(r['total_ban'], 0)} | **Tổng views:** {fmt(r['total_views'], 0)} | **Views/Tin:** {fmt(r['views_tin'], 1)}",
            f"- **MFV:** {fmt(r['mfv'], 0)}",
            f"- **Giá:** {fmt(r['price'], 1)} tr/m² | **Gap vs {benchmark}:** {fmt(r['gia_gap_pct'], 0)}%",
            f"- **Cung-Cầu:** {fmt(r['cung_cau'], 2)} ({cc_label(r['cung_cau'])})",
            f"- **Cắt lỗ:** {fmt(r['distress_ratio'], 1)}% | **Yield:** {fmt(r['rental_yield'], 2)}%",
            "",
        ])

    lines.extend([
        "---", "",
        "## IV. GIẢI THÍCH CHỈ SỐ PHÁI SINH", "",
        "| Chỉ số | Ý nghĩa | Công thức |",
        "|--------|---------|-----------|",
        "| **Cycle Index** (0-100) | Vị trí trong chu kỳ BĐS | Cắt lỗ + Cung-Cầu + YoY + Absorption + Views |",
        "| **Heat Score** | Mức \"nóng\" tổng hợp | MFV(40%) + Views(25%) + C-C(15%) + YoY(15%) + Health(5%) |",
        "| **MFV** | Tốc độ dòng tiền | Tổng tin × Absorption × Views × (1 + YoY) |",
        "| **MFSI** | Hướng dòng tiền | Thanh khoản(40%) + Yield(30%) + Gap(20%) + Tâm lý(10%) |",
        "| **Cung-Cầu** | Cân bằng thị trường | Views/Tin normalized (5 views = 1.0) |",
        "| **Giá gap** | Chênh lệch vs benchmark | (Giá - Benchmark) / Benchmark × 100% |",
        "",
        "> ⚠️ **Views/Tin** và **Giá tăng YoY** lấy từ `chi_so_config.json` (cập nhật thủ công).",
        "> Các chỉ số khác tính tự động từ dữ liệu crawl.",
        "",
        "---", "",
        f"📅 **Tính lúc:** {timestamp} | 🤖 **Script:** tinh_chi_so_phai_sinh.py",
    ])

    return "\n".join(lines)


def save_history(results: dict, filepath: Path):
    """Append lịch sử chỉ số phái sinh."""
    headers = [
        "Ngày", "Khu vực",
        "Tổng tin", "Tổng views", "Views/Tin",
        "Cycle", "Heat Score", "MFV", "MFSI",
        "Cung-Cầu", "Giá gap (%)",
        "Giá (tr/m²)", "Cắt lỗ (%)", "YoY (%)",
    ]
    today = datetime.now().strftime("%Y-%m-%d")

    # Chuẩn bị rows
    rows = []
    for name, r in results.items():
        rows.append([
            today, name,
            int(r.get("total_ban", 0)),
            int(r.get("total_views", 0)),
            r.get("views_tin", ""),
            r.get("cycle", ""),
            r.get("heat_score", ""),
            r.get("mfv", ""),
            round(r["mfsi"], 1) if r.get("mfsi") else "",
            round(r["cung_cau"], 2) if r.get("cung_cau") else "",
            round(r["gia_gap_pct"], 1) if r.get("gia_gap_pct") is not None else "",
            r.get("price", ""),
            r.get("distress_ratio", ""),
            r.get("gia_yoy", ""),
        ])

    # DC01: Chống ghi trùng
    if HAS_DEFENSIVE:
        safe_append_csv(filepath, rows, headers, date_column="Ngày")
    else:
        file_exists = filepath.exists()
        try:
            with open(filepath, "a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(headers)
                writer.writerows(rows)
        except PermissionError:
            print(f"  !! Không ghi được {filepath.name} — đóng file trong Excel/Editor rồi chạy lại.")


def save_excel(results: dict, filepath: Path):
    """Tạo file Excel: Sheet 1 = Lịch sử theo ngày, Sheet 2 = Biểu đồ đường xu hướng."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  !! Cần cài openpyxl: pip install openpyxl")
        return

    # Đọc lịch sử từ CSV
    history_path = OUTPUT_DIR / "lich_su_chi_so_phai_sinh.csv"
    history_rows = []
    if history_path.exists():
        with open(history_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                history_rows.append(row)

    if not history_rows:
        print("  !! Chưa có dữ liệu lịch sử. Cần chạy ít nhất 1 lần.")
        return

    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_odd = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ============================================================
    # SHEET 1: BẢNG LỊCH SỬ THEO NGÀY (ưu tiên Tin, Views, Views/Tin)
    # ============================================================
    ws1 = wb.active
    ws1.title = "Lịch sử theo ngày"

    # Sắp xếp lại thứ tự cột: ưu tiên Tin, Views, Views/Tin
    priority_cols = ["Ngày", "Khu vực", "Tổng tin", "Tổng views", "Views/Tin", "Giá (tr/m²)"]
    all_csv_headers = list(history_rows[0].keys())
    remaining = [h for h in all_csv_headers if h not in priority_cols]
    ordered_headers = priority_cols + remaining

    # Màu highlight cho 3 cột ưu tiên
    priority_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col, h in enumerate(ordered_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = priority_fill if h in ("Tổng tin", "Tổng views", "Views/Tin", "Giá (tr/m²)") else header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(history_rows, 2):
        fill = fill_odd if row_idx % 2 == 0 else fill_even
        for col_idx, key in enumerate(ordered_headers, 1):
            val = row.get(key, "")
            try:
                if "." in str(val):
                    val = float(val)
                elif val and str(val).isdigit():
                    val = int(val)
            except (ValueError, AttributeError):
                pass
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = name_align if col_idx <= 2 else data_align
            cell.fill = fill
            cell.border = thin_border

    # Chiều rộng cột
    for i, h in enumerate(ordered_headers, 1):
        w = {"Ngày": 12, "Khu vực": 16, "Tổng tin": 10, "Tổng views": 12, "Views/Tin": 10, "Giá (tr/m²)": 12}.get(h, 11)
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.freeze_panes = "C2"

    # ============================================================
    # SHEET 2: 11 BIỂU ĐỒ — MỖI KHU VỰC 1 BIỂU ĐỒ
    # ============================================================
    dates = []
    regions = []
    seen_dates = set()
    seen_regions = set()

    for row in history_rows:
        d = row.get("Ngày", "")
        r = row.get("Khu vực", "")
        if d and d not in seen_dates:
            dates.append(d)
            seen_dates.add(d)
        if r and r not in seen_regions:
            regions.append(r)
            seen_regions.add(r)

    # Build lookup: (ngày, khu vực) → {tin, views, views_tin, gia}
    lookup = {}
    for row in history_rows:
        key = (row.get("Ngày", ""), row.get("Khu vực", ""))
        lookup[key] = {
            "tin": _float(row.get("Tổng tin")),
            "views": _float(row.get("Tổng views")),
            "views_tin": _float(row.get("Views/Tin")),
            "gia": _float(row.get("Giá (tr/m²)")),
        }

    # ============================================================
    # SHEET 2+: MỖI KHU VỰC 1 SHEET RIÊNG
    # ============================================================
    dark_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

    for region_idx, region in enumerate(regions):
        # Tạo sheet mới cho khu vực (tên sheet max 31 ký tự)
        sheet_name = region[:31]
        ws = wb.create_sheet(sheet_name)

        # Header: Ngày | Tổng tin | Tổng views | Views/Tin | Giá (tr/m²)
        for col, h in enumerate(["Ngày", "Tổng tin", "Tổng views", "Views/Tin", "Giá (tr/m²)"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = header_align

        # Data: mỗi dòng = 1 ngày
        for row_offset, date in enumerate(dates, 1):
            ws.cell(row=1 + row_offset, column=1, value=date)
            data = lookup.get((date, region), {})
            ws.cell(row=1 + row_offset, column=2, value=data.get("tin"))
            ws.cell(row=1 + row_offset, column=3, value=data.get("views"))
            ws.cell(row=1 + row_offset, column=4, value=data.get("views_tin"))
            ws.cell(row=1 + row_offset, column=5, value=data.get("gia"))

        last_row = 1 + len(dates)

        # Chiều rộng cột
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12

        # --- Tạo Line Chart ---
        from openpyxl.chart import LineChart, Reference

        chart = LineChart()
        chart.title = region
        chart.y_axis.title = "Tin / Views"
        chart.x_axis.title = "Ngày"
        chart.width = 26
        chart.height = 14
        chart.style = 10

        # Tổng tin (cột 2) + Tổng views (cột 3)
        data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_row)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Style: Tổng tin = xanh dương, Tổng views = cam
        if len(chart.series) > 0:
            chart.series[0].graphicalProperties.line.solidFill = "2F5496"
            chart.series[0].graphicalProperties.line.width = 25000
        if len(chart.series) > 1:
            chart.series[1].graphicalProperties.line.solidFill = "ED7D31"
            chart.series[1].graphicalProperties.line.width = 25000

        # Views/Tin + Giá trên trục Y phụ
        chart2 = LineChart()
        vt_ref = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=last_row)
        chart2.add_data(vt_ref, titles_from_data=True)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = "Views/Tin — Giá (tr/m²)"

        # Style: Views/Tin = xanh lá nét đứt, Giá = đỏ nét liền
        if len(chart2.series) > 0:
            s = chart2.series[0]
            s.graphicalProperties.line.solidFill = "70AD47"
            s.graphicalProperties.line.width = 25000
            s.graphicalProperties.line.dashStyle = "dash"
        if len(chart2.series) > 1:
            s = chart2.series[1]
            s.graphicalProperties.line.solidFill = "E63946"
            s.graphicalProperties.line.width = 25000

        chart.y_axis.crosses = "min"
        chart += chart2

        ws.add_chart(chart, "A" + str(last_row + 2))


    # Save
    try:
        wb.save(filepath)
    except PermissionError:
        print(f"  !! Không ghi được {filepath.name} — đóng file Excel rồi chạy lại.")


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Tính chỉ số phái sinh thị trường BĐS")
    parser.add_argument(
        "--only", type=str, default="",
        help="Chỉ tính cho các khu vực cụ thể (cách nhau bằng dấu phẩy)"
    )
    args = parser.parse_args()

    print("=" * 60)
    print("TÍNH CHỈ SỐ PHÁI SINH THỊ TRƯỜNG BĐS")
    print("=" * 60)

    # 1. Đọc dữ liệu crawl
    crawl_path = OUTPUT_DIR / "lich_su_chi_so.csv"
    print(f"\n1. Đọc dữ liệu crawl: {crawl_path}")
    crawl_data = load_crawl_data(crawl_path)
    if not crawl_data:
        print("!! Không có dữ liệu crawl. Chạy crawl_chi_so_thi_truong.py trước.")
        return
    print(f"   → {len(crawl_data)} khu vực")

    # 2. Đọc config
    config_path = OUTPUT_DIR / "chi_so_config.json"
    print(f"\n2. Đọc config: {config_path}")
    config = load_config(config_path)
    if not config:
        print("!! Sử dụng giá trị mặc định cho Views/Tin và YoY.")
        config = {"views_per_tin": {}, "gia_yoy_pct": {}, "benchmark_khu_vuc": "Nam Từ Liêm"}

    # DC06: Kiểm tra config có quá hạn không
    if HAS_DEFENSIVE:
        check_config_freshness(config_path)

    # 3. Lọc khu vực nếu có --only
    if args.only:
        only_list = [x.strip() for x in args.only.split(",")]
        crawl_data = {k: v for k, v in crawl_data.items()
                      if any(o.lower() in k.lower() for o in only_list)}
        print(f"\n   → Lọc: {len(crawl_data)} khu vực")

    # 4. Tính chỉ số phái sinh
    print(f"\n3. Tính chỉ số phái sinh...")
    results = calculate_derived(crawl_data, config)

    # 5. In tóm tắt
    print(f"\n{'=' * 80}")
    print(f"{'Khu vực':<16} {'Tin':>6} {'Views':>8} {'V/Tin':>6} {'Cycle':>6} {'Heat':>6} {'MFV':>8} {'C-C':>6} {'Gap%':>6}")
    print(f"{'-' * 16} {'-' * 6} {'-' * 8} {'-' * 6} {'-' * 6} {'-' * 6} {'-' * 8} {'-' * 6} {'-' * 6}")
    for name, r in sorted(results.items(), key=lambda x: x[1].get("heat_score", 0), reverse=True):
        gap = f"{r['gia_gap_pct']:+.0f}" if r['gia_gap_pct'] is not None else "—"
        print(f"{name:<16} {r['total_ban']:>6.0f} {r['total_views']:>8.0f} {r['views_tin']:>6.1f} "
              f"{r['cycle']:>6.0f} {r['heat_score']:>6.0f} {r['mfv']:>8.0f} "
              f"{r['cung_cau']:>6.2f} {gap:>6}")

    # 6. Xuất báo cáo
    md_path = OUTPUT_DIR / "CHI_SO_PHAI_SINH.md"
    report = generate_report(results, crawl_data, config)
    md_path.write_text(report, encoding="utf-8")
    print(f"\n4. Báo cáo MD: {md_path}")

    # 7. Xuất Excel (Sheet 1 = data, Sheet 2 = biểu đồ)
    excel_path = OUTPUT_DIR / "CHI_SO_PHAI_SINH.xlsx"
    save_excel(results, excel_path)
    print(f"5. Excel + Biểu đồ: {excel_path}")

    # 8. Lưu lịch sử
    history_path = OUTPUT_DIR / "lich_su_chi_so_phai_sinh.csv"
    save_history(results, history_path)
    print(f"6. Lịch sử: {history_path}")

    print(f"\n{'=' * 60}")
    print(f"HOÀN THÀNH! {len(results)} khu vực đã tính.")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
